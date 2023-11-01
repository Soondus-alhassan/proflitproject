from flask import Flask, render_template, request, redirect, url_for, send_file, jsonify
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import requests
import numpy as np
import re
import phonenumbers
import upload_file


app = Flask(__name__)

UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
COUNTRIES = ["USA", "United Kingdom", "Canada"]

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'xlsx'}

def validate_uk_postcode(postcode):
    url = f"https://api.postcodes.io/postcodes/{postcode}/validate"
    response = requests.get(url)
    if response.status_code == 200:
        data = response.json()
        return data.get("result", False)
    return False

def detect_country_column(df):
    for col in df.columns:
        unique_values = df[col].unique()
        country_matches = sum([1 for val in unique_values if str(val) in COUNTRIES])
        if country_matches > len(unique_values) * 0.5:  # if more than 50% match
            return col
    return None

def standardize_country_name(country_name):
    if country_name in ["U.S.A.", "United States"]:
        return "USA"
    return country_name

@app.route('/')
def index():
    return render_template('original.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return redirect(request.url)
    
    file = request.files['file']

    if file.filename == '':
        return redirect(request.url)

    if file and allowed_file(file.filename):
        new_filename = 'sanitized_Filename.xlsx'
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], new_filename)
        file.save(file_path)
        
        return redirect(url_for('process_file', filename=new_filename))
    else:
        return 'Invalid file format. Please upload an Excel file with .xlsx extension.'

@app.route('/process/<filename>')
def process_file(filename):
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    workbook = load_workbook(file_path)
    worksheet = workbook.active
    company_names = set()

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        company_col, postcode_col = row[0], row[1]
        if company_col.value in company_names:
            company_col.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        else:
            company_names.add(company_col.value)

        if not validate_uk_postcode(postcode_col.value):
            postcode_col.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    df = pd.read_excel(file_path)

    country_col = detect_country_column(df)
    if country_col:
        df[country_col] = df[country_col].apply(standardize_country_name)

    df.to_excel(file_path, index=False)
    workbook.save(file_path)
    return send_file(file_path, as_attachment=True, download_name=filename)


@app.route('/data-profiling', methods=['GET', 'POST'])
def data_profiling():
    if request.method == 'POST':
        file = request.files.get('file')
        if file and allowed_file(file.filename):
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
            file.save(file_path)
            df = pd.read_excel(file_path)
            profile = {
                col: {
                    'Total records': len(df[col]),
                    'Unique values': df[col].nunique(),
                    'Missing values': df[col].isna().sum(),
                    'Top 5 values': dict(df[col].value_counts().head(5))
                } for col in df.columns
            }
            return render_template('data_profiling_results.html', profile=profile)
    return render_template('data_profiling.html')

def perform_data_profiling(file_path):
    df = pd.read_excel(file_path)
    profile = {}

    for col in df.columns:
        profile[col] = {
            'Total records': int(len(df[col])),
            'Unique values': int(df[col].nunique()),
            'Missing values': int(df[col].isna().sum()),
            'Top 5 values': {k: int(v) if isinstance(v, np.int64) else v for k, v in df[col].value_counts().head(5).to_dict().items()}
        }
    return profile
    v
def dict_to_string(d):
    return ', '.join([f"{k}: {v}" for k, v in d.items()])

app.jinja_env.filters['to_string'] = dict_to_string


@app.route('/validate', methods=['POST'])
def validate_data():
    if 'file' not in request.files:
        return redirect(url_for('index'))
    
    file = request.files['file']

    if file.filename == '':
        return redirect(url_for('index'))

    if file and allowed_file(file.filename):
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(file_path)
        
        # Here you can add the logic to validate the data.
        # For now, let's assume `perform_validate` is a function that 
        # returns a list of issues with the data.
        issues = perform_validate(file_path)
        
        return render_template('validate.html', issues=issues)

    else:
        return redirect(url_for('index'))

def validate_country_names(df, country_col):
    invalid_countries = df[~df[country_col].isin(COUNTRIES)]
    issues = []
    for index, row in invalid_countries.iterrows():
        issues.append({
            "type": "Invalid Country",
            "column_name": country_col,
            "row_number": index + 2,  # +1 for 0-based index, +1 for header row
            "details": f"Invalid country name: {row[country_col]}"
        })
    return issues
def perform_validate(file_path):
    df = pd.read_excel(file_path)
    issues = []

    # Check if any values in a column named "ExampleColumn" are null
    if "ExampleColumn" in df.columns and df["ExampleColumn"].isnull().any():
        issues.append({
            "type": "Missing Value",
            "column_name": "ExampleColumn",
            "row_number": df[df["ExampleColumn"].isnull()].index[0] + 2,
            "details": "Value is missing."
        })

    # Country name validation
    country_col = detect_country_column(df)
    if country_col:
        issues.extend(validate_country_names(df, country_col))

    return issues



if __name__ == '__main__':
    app.run(debug=True)
