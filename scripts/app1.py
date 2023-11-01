from flask import Flask, render_template, request, send_file
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import phonenumbers
import os
import re

app = Flask(__name__)

VALIDATED_FOLDER = 'validated_data'
if not os.path.exists(VALIDATED_FOLDER):
    os.makedirs(VALIDATED_FOLDER)


UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
blue_fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() == 'xlsx'

def check_mandatory(value):
    return bool(value)

def check_uniqueness(value, col_values):
    return col_values.count(value) <= 1

def check_if_valid_phonenumber(value):
    try:
        phonenumbers.parse(value)
        return True
    except:
        return False

def check_valid_values(value, valid_values=[0, 1, 2]):
    return value in valid_values

def check_max_length(value, max_len=13):
    return len(value) <= max_len

def check_if_valid_email(value):
    email_regex = r"[^@]+@[^@]+\.[^@]+"
    return re.match(email_regex, value)

def check_if_alphabetic(value):
    return value.isalpha()

def check_min_length(value, min_len=5):
    return len(value) >= min_len

def check_if_numeric(value):
    if isinstance(value, int):
        return True
    elif isinstance(value, str):
        return value.isnumeric()
    else:
        return False

def check_mandatory_for_another(value, another_value):
    return bool(value) or not another_value
def validate_excel(file_path):
    workbook = load_workbook(file_path)
    worksheet = workbook.active

    max_column = worksheet.max_column
    error_col = max_column + 1
    worksheet.cell(row=1, column=error_col, value="Error Messages")

    col_headers = {}
    for cell in worksheet[1]:  # Assuming the first row contains headers
        col_headers[cell.column_letter] = cell.value

    col_values = {header: [] for header in col_headers.values()}

    for column_letter, header in col_headers.items():
     col_values[header] = [cell.value for cell in worksheet[column_letter][1:]]  # Exclude header cell

    # (The rest of your validate_excel function remains the same...)

    # Iterate over each row (excluding the header)
    for row_cells in worksheet.iter_rows(min_row=2):
        error_messages = []  # Reset the error messages for every new row

        for cell in row_cells:
            attribute = col_headers[cell.column_letter]
            value = cell.value

            # Full Name validation
            if attribute == "full_name":
                if not check_mandatory(value):
                    cell.fill = red_fill
                    error_messages.append("Full Name: This is mandatory")

            # Phone validation
            elif attribute == "phone":
                if not check_mandatory(value):
                    cell.fill = red_fill
                    error_messages.append("Phone: This is mandatory")
                elif not check_if_valid_phonenumber(value):
                    cell.fill = red_fill
                    error_messages.append("Phone: Invalid phone number")
                elif not check_uniqueness(value, col_values["phone"]):
                    cell.fill = red_fill
                    error_messages.append("Phone: Duplicate phone number")
                elif not check_max_length(value):
                    cell.fill = red_fill
                    error_messages.append("Phone: Exceeded Max Length")
                elif not check_min_length(value):
                    cell.fill = red_fill
                    error_messages.append("Phone: Below Min Length")

            # Street Address validation
            elif attribute == "street_address#1":
                if not check_mandatory(value):
                    cell.fill = red_fill
                    error_messages.append("Street Address: This is mandatory")

            # Marketing Consent validation
            elif attribute == "marketing_consent":
                if not check_mandatory(value):
                    cell.fill = red_fill
                    error_messages.append("Marketing Consent: This is mandatory")
                elif not check_if_numeric(value):
                    cell.fill = red_fill
                    error_messages.append("Marketing Consent: Must be a number")
                elif not check_valid_values(value):
                    cell.fill = red_fill
                    error_messages.append("Marketing Consent: Invalid value")

            # Email validation
            elif attribute == "email":
                if not check_if_valid_email(value):
                    cell.fill = red_fill
                    error_messages.append("Email: Invalid email format")

            # City validation
            elif attribute == "city":
                if not check_mandatory_for_another(value, col_values.get("street_address#1")):
                    cell.fill = red_fill
                    error_messages.append("City: Mandatory if street address is provided")
                elif not check_if_alphabetic(value):
                    cell.fill = red_fill
                    error_messages.append("City: Must be alphabetic")

        # After checking all cells for a row, write the error messages
        if error_messages:
            worksheet.cell(row=row_cells[0].row, column=error_col).value = ', '.join(error_messages)

    # Save workbook with highlighted errors
    save_path = os.path.join(app.config['UPLOAD_FOLDER'], "validated_" + os.path.basename(file_path))
    workbook.save(save_path)
    # Save workbook with highlighted errors
    save_path = os.path.join(VALIDATED_FOLDER, "validated_" + os.path.basename(file_path))
    workbook.save(save_path)
    return save_path


#return save_path



@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        file = request.files.get('file')
        if file and allowed_file(file.filename):
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
            file.save(file_path)
            validated_file_path = validate_excel(file_path)
            return send_file(validated_file_path, as_attachment=True, download_name="validated_" + file.filename)

    return render_template('upload.html')



if __name__ == '__main__':
    app.run(debug=True)